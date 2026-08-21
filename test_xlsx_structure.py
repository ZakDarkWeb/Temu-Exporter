from pathlib import Path
from openpyxl import load_workbook

path = Path(__file__).with_name('test-output.xlsx')
wb = load_workbook(path)
assert wb.sheetnames == ['Orders', 'Extraction Status']
ws = wb['Orders']
assert [cell.value for cell in ws[1]] == [
    'Shipping Date', 'Order Date', 'Tracking Number', 'Order No', 'Customer Name',
    'Product Details', 'Qty (No)', 'Est. Revenue', 'Shipping Cost'
]
assert ws['A2'].value == 'Aug 20, 2026'
assert ws['B2'].value == 'Aug 19, 2026'
assert ws['G2'].value == 1 and ws['G2'].data_type == 'n' and ws['G2'].number_format == '0'
assert ws['H2'].value == 16.02 and ws['H2'].data_type == 'n'
assert ws['I2'].value == 6.24 and ws['I2'].data_type == 'n'
assert ws['D2'].value == ws['D3'].value == 'PO-211-01861395087993272'
assert ws['F3'].value == 'TaylorMade Blue Ink SpeedSoft Golf Balls'
assert ws['G3'].value == 1 and ws['G3'].number_format == '0'
assert ws['H3'].value is None and ws['I3'].value is None
assert ws.freeze_panes == 'A2'
assert ws.auto_filter.ref == 'A1:I3'
assert ws['A1'].fill.fgColor.rgb == 'FF00B050'
assert ws['A1'].font.bold is True
assert ws['A1'].font.color.type == 'rgb' and ws['A1'].font.color.rgb == 'FFFFFFFF'
assert ws.tables['TemuOrders'].ref == 'A1:I3'
status = wb['Extraction Status']
assert status['A1'].value == 'Time'
assert status['E2'].value == 'Temu opened a no-auth page.'
print('xlsx structure: PASS')
print('sheets:', wb.sheetnames)
print('orders headers:', [cell.value for cell in ws[1]])
print('numeric cells:', ws['G2'].value, ws['H2'].value, ws['I2'].value)
print('table ref:', ws.tables['TemuOrders'].ref)
